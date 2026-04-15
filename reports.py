from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import send_file
from models import db, FinanceTransaction, Order, Employee
from sqlalchemy import func

def generate_finance_report_pdf(start_date, end_date):
    """Генерация PDF отчёта по финансам за период"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, spaceAfter=20)
    elements = []
    
    # Заголовок
    elements.append(Paragraph(f"Финансовый отчёт за {start_date} – {end_date}", title_style))
    elements.append(Spacer(1, 10))
    
    # Данные из БД
    transactions = FinanceTransaction.query.filter(
        FinanceTransaction.date >= start_date,
        FinanceTransaction.date <= end_date
    ).order_by(FinanceTransaction.date).all()
    
    income_total = sum(t.amount for t in transactions if t.type == 'income')
    expense_total = sum(t.amount for t in transactions if t.type == 'expense')
    profit = income_total - expense_total
    
    # Итоговая строка
    summary_data = [
        ['Итого доходов:', f'{income_total:,.2f} руб.'],
        ['Итого расходов:', f'{expense_total:,.2f} руб.'],
        ['Прибыль:', f'{profit:,.2f} руб.']
    ]
    summary_table = Table(summary_data, colWidths=[100, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Таблица транзакций
    data = [['Дата', 'Тип', 'Категория', 'Сумма, руб.', 'Описание']]
    for t in transactions:
        data.append([
            t.date.strftime('%d.%m.%Y %H:%M'),
            'Доход' if t.type == 'income' else 'Расход',
            t.category,
            f'{t.amount:,.2f}',
            t.description or ''
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'finance_report_{start_date}_{end_date}.pdf', mimetype='application/pdf')

def generate_finance_report_excel(start_date, end_date):
    """Генерация Excel отчёта по финансам"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансы"
    
    # Заголовки
    headers = ['Дата', 'Тип', 'Категория', 'Сумма', 'Описание']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    transactions = FinanceTransaction.query.filter(
        FinanceTransaction.date >= start_date,
        FinanceTransaction.date <= end_date
    ).order_by(FinanceTransaction.date).all()
    
    income_total = expense_total = 0
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=2, value='Доход' if t.type == 'income' else 'Расход')
        ws.cell(row=row, column=3, value=t.category)
        ws.cell(row=row, column=4, value=t.amount)
        ws.cell(row=row, column=5, value=t.description or '')
        if t.type == 'income':
            income_total += t.amount
        else:
            expense_total += t.amount
    
    profit = income_total - expense_total
    row = len(transactions) + 2
    ws.cell(row=row, column=3, value="Итого доходов:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=income_total)
    row += 1
    ws.cell(row=row, column=3, value="Итого расходов:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=expense_total)
    row += 1
    ws.cell(row=row, column=3, value="Прибыль:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=profit)
    
    for col in [1,2,3,4,5]:
        ws.column_dimensions[chr(64+col)].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'finance_report_{start_date}_{end_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def generate_salary_report_pdf(month_year):
    """Зарплатная ведомость за месяц (формат YYYY-MM)"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"Зарплатная ведомость за {month_year}", styles['Title']))
    elements.append(Spacer(1, 10))
    
    # Получаем зарплатные транзакции за месяц
    start_date = datetime.strptime(f"{month_year}-01", "%Y-%m-%d")
    if month_year.endswith('-12'):
        end_date = datetime(start_date.year, 12, 31)
    else:
        end_date = datetime(start_date.year, start_date.month+1, 1) - timedelta(days=1)
    
    salary_trans = FinanceTransaction.query.filter(
        FinanceTransaction.category == 'Зарплата',
        FinanceTransaction.date >= start_date,
        FinanceTransaction.date <= end_date
    ).all()
    
    data = [['Сотрудник', 'Сумма, руб.', 'Дата выплаты']]
    total = 0
    for t in salary_trans:
        # Из описания пытаемся вытащить имя сотрудника
        emp_name = t.description.replace('Выплата ', '') if t.description else '—'
        data.append([emp_name, f'{t.amount:,.2f}', t.date.strftime('%d.%m.%Y')])
        total += t.amount
    data.append(['ИТОГО:', f'{total:,.2f}', ''])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'salary_{month_year}.pdf', mimetype='application/pdf')

def generate_orders_report_excel(start_date, end_date):
    """Отчёт по заказам (прибыль, загруженность мастеров)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    headers = ['ID', 'Клиент', 'Модель', 'Дата приёма', 'Дата завершения', 'Статус', 'Цена, руб.', 'Мастер']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    orders = Order.query.filter(Order.start_time >= start_date, Order.start_time <= end_date).all()
    total_revenue = 0
    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.id)
        ws.cell(row=row, column=2, value=o.customer_name)
        ws.cell(row=row, column=3, value=o.device_model)
        ws.cell(row=row, column=4, value=o.start_time.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=5, value=o.completed_at.strftime('%d.%m.%Y') if o.completed_at else '—')
        ws.cell(row=row, column=6, value=o.status)
        ws.cell(row=row, column=7, value=o.price)
        master = Employee.query.get(o.responsible_employee_id)
        ws.cell(row=row, column=8, value=master.full_name if master else '—')
        if o.status == 'completed':
            total_revenue += o.price
    
    row = len(orders) + 2
    ws.cell(row=row, column=6, value="Общая выручка:").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_revenue)
    
    for col in range(1, 9):
        ws.column_dimensions[chr(64+col)].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'orders_report_{start_date}_{end_date}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')