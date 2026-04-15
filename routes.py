from functools import wraps
from threading import Thread
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import db, WarehouseItem, Employee, Order, FinanceTransaction, BlacklistClient, PriceItem, RecurringPayment, User, Notification, OrderLog, WarrantyCard
from datetime import datetime, timedelta
from sqlalchemy import func
from utils import check_deadlines_and_notify, send_order_ready_email_with_act, generate_act_pdf_buffer, log_order_change
import json

main_bp = Blueprint('main', __name__)

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.has_role(*roles):
                flash('Доступ запрещён', 'danger')
                return redirect(url_for('main.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def add_transaction(type_, category, amount, description):
    tr = FinanceTransaction(type=type_, category=category, amount=amount, description=description)
    db.session.add(tr)
    db.session.commit()

def recalc_balance():
    inc = db.session.query(func.sum(FinanceTransaction.amount)).filter_by(type='income').scalar() or 0
    exp = db.session.query(func.sum(FinanceTransaction.amount)).filter_by(type='expense').scalar() or 0
    return inc - exp

# ------------------ Главная страница ------------------
@main_bp.route('/')
@login_required
def index():
    Thread(target=check_deadlines_and_notify).start()
    return render_template('index.html')

# ------------------ Страницы ------------------
@main_bp.route('/warehouse')
@role_required('admin', 'manager')
def warehouse_page():
    return render_template('warehouse.html')

@main_bp.route('/employees')
@role_required('admin', 'manager')
def employees_page():
    return render_template('employees.html')

@main_bp.route('/orders_page')
@login_required
def orders_page():
    return render_template('orders.html')

@main_bp.route('/finance')
@role_required('admin', 'accountant', 'manager')
def finance_page():
    return render_template('finance.html')

@main_bp.route('/blacklist_page')
@role_required('admin', 'manager')
def blacklist_page():
    return render_template('blacklist.html')

@main_bp.route('/pricelist_page')
@login_required
def pricelist_page():
    return render_template('pricelist.html')

@main_bp.route('/calendar_page')
@login_required
def calendar_page():
    return render_template('calendar.html')

@main_bp.route('/settings')
@login_required
def settings_page():
    return render_template('settings.html')

@main_bp.route('/users')
@role_required('admin')
def users_list():
    users = User.query.all()
    return render_template('users.html', users=users)

@main_bp.route('/warranty')
@login_required
def warranty_page():
    return render_template('warranty.html')

# ------------------ API: Дашборд и уведомления ------------------
@main_bp.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    active_orders = Order.query.filter(Order.status != 'completed').count()
    overdue_orders = Order.query.filter(Order.deadline < datetime.utcnow(), Order.status != 'completed').count()
    today_orders = Order.query.filter(func.date(Order.start_time) == datetime.utcnow().date()).count()
    balance = recalc_balance()
    low_stock = WarehouseItem.query.filter(WarehouseItem.quantity < 3).count()
    return jsonify({
        'active_orders': active_orders,
        'overdue_orders': overdue_orders,
        'today_orders': today_orders,
        'balance': balance,
        'low_stock': low_stock
    })

@main_bp.route('/api/notifications')
@login_required
def get_notifications():
    role = current_user.role
    notif = Notification.query.filter(
        (Notification.user_role.like(f'%{role}%')) | (Notification.user_id == current_user.id),
        Notification.is_read == False
    ).order_by(Notification.created_at.desc()).limit(20).all()
    return jsonify([{
        'id': n.id,
        'title': n.title,
        'message': n.message,
        'link': n.link,
        'created_at': n.created_at.isoformat()
    } for n in notif])

@main_bp.route('/api/notifications/<int:id>/read', methods=['POST'])
@login_required
def mark_notification_read(id):
    n = Notification.query.get_or_404(id)
    n.is_read = True
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_read():
    role = current_user.role
    Notification.query.filter(
        (Notification.user_role.like(f'%{role}%')) | (Notification.user_id == current_user.id),
        Notification.is_read == False
    ).update({'is_read': True})
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Склад ------------------
@main_bp.route('/api/warehouse', methods=['GET'])
@role_required('admin', 'manager')
def get_warehouse():
    items = WarehouseItem.query.all()
    return jsonify([{
        'id': i.id, 'name': i.name, 'quantity': i.quantity,
        'weight': i.weight, 'size': i.size, 'cost_price': i.cost_price,
        'price_with_markup': round(i.cost_price * 1.2, 2)
    } for i in items])

@main_bp.route('/api/warehouse', methods=['POST'])
@role_required('admin', 'manager')
def add_warehouse():
    data = request.json
    item = WarehouseItem(
        name=data['name'],
        quantity=data['quantity'],
        weight=data.get('weight', ''),
        size=data.get('size', ''),
        cost_price=data['cost_price']
    )
    db.session.add(item)
    add_transaction('expense', 'Покупка товара', data['cost_price'], f'Закупка: {data["name"]}')
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/warehouse/<int:id>', methods=['DELETE'])
@role_required('admin', 'manager')
def delete_warehouse(id):
    item = WarehouseItem.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Сотрудники ------------------
@main_bp.route('/api/employees', methods=['GET'])
@role_required('admin', 'manager')
def get_employees():
    emps = Employee.query.filter_by(fired=False).all()
    return jsonify([{
        'id': e.id, 'full_name': e.full_name, 'passport': e.passport,
        'details': e.details, 'position': e.position, 'salary_value': e.salary_value
    } for e in emps])

@main_bp.route('/api/employees', methods=['POST'])
@role_required('admin', 'manager')
def add_employee():
    data = request.json
    emp = Employee(
        full_name=data['full_name'],
        passport=data.get('passport', ''),
        details=data.get('details', ''),
        position=data['position'],
        salary_value=data['salary_value']
    )
    db.session.add(emp)
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/employees/<int:id>/fire', methods=['POST'])
@role_required('admin', 'manager')
def fire_employee(id):
    emp = Employee.query.get_or_404(id)
    emp.fired = True
    emp.fire_reason = request.json.get('reason', '')
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Заказы ------------------
@main_bp.route('/api/orders', methods=['POST'])
@role_required('admin', 'receiver', 'manager')
def create_order():
    data = request.json
    blacklisted = BlacklistClient.query.filter(
        (BlacklistClient.full_name == data['customer_name']) | 
        (BlacklistClient.phone == data['phone'])
    ).first()
    if blacklisted:
        return jsonify({'error': f'Клиент в чёрном списке. Причина: {blacklisted.reason}'}), 400
    order = Order(
        customer_name=data['customer_name'],
        phone=data['phone'],
        device_model=data['device_model'],
        serial_number=data.get('serial_number', ''),
        imei=data.get('imei', ''),
        main_problem=data.get('main_problem', ''),
        detected_problem=data.get('detected_problem', ''),
        price=data['price'],
        deadline=datetime.fromisoformat(data['deadline']),
        status='waiting_parts' if data.get('waiting_parts') else 'in_progress',
        responsible_employee_id=data['responsible_employee_id']
    )
    db.session.add(order)
    db.session.commit()
    add_transaction('income', 'Прием заказа', data['price'], f'Заказ №{order.id}')
    log_order_change(
        order_id=order.id,
        user_id=current_user.id,
        username=current_user.username,
        action='create',
        comment=f'Создан заказ на сумму {order.price} руб., мастер id={order.responsible_employee_id}'
    )
    return jsonify({'status': 'ok', 'id': order.id})

@main_bp.route('/api/orders/<int:id>', methods=['PUT'])
@role_required('admin', 'manager', 'receiver')
def update_order(id):
    order = Order.query.get_or_404(id)
    data = request.json
    order.customer_name = data.get('customer_name', order.customer_name)
    order.phone = data.get('phone', order.phone)
    order.device_model = data.get('device_model', order.device_model)
    order.main_problem = data.get('main_problem', order.main_problem)
    order.detected_problem = data.get('detected_problem', order.detected_problem)
    order.price = data.get('price', order.price)
    if 'deadline' in data:
        order.deadline = datetime.fromisoformat(data['deadline'])
    order.status = data.get('status', order.status)
    order.responsible_employee_id = data.get('responsible_employee_id', order.responsible_employee_id)
    db.session.commit()
    log_order_change(
        order_id=order.id,
        user_id=current_user.id,
        username=current_user.username,
        action='edit',
        comment='Изменены данные заказа'
    )
    return jsonify({'status': 'ok'})

@main_bp.route('/api/orders/<int:id>/complete', methods=['POST'])
@role_required('admin', 'manager', 'master')
def complete_order(id):
    order = Order.query.get_or_404(id)
    if order.status == 'completed':
        return jsonify({'error': 'Already completed'}), 400
    order.status = 'completed'
    order.completed_at = datetime.utcnow()
    master_percent = request.json.get('master_percent', 67)
    bonus_days = request.json.get('bonus_days', 3)
    bonus_percent = request.json.get('bonus_percent', 10)
    days_taken = (order.completed_at - order.start_time).total_seconds() / 86400
    base_salary = order.price * master_percent / 100
    factor = 1
    if days_taken < bonus_days:
        factor = 1 + bonus_percent/100
    elif days_taken > bonus_days:
        factor = 1 - bonus_percent/100
    final_salary = round(base_salary * factor, 2)
    add_transaction('expense', 'ЗП мастеру (ремонт)', final_salary, f'Заказ {order.id}')
    db.session.commit()
    # Создание гарантийных талонов
    used_parts = request.json.get('used_parts', False)
    WarrantyCard.create_for_order(order, 'work', f'Гарантия на выполненные работы по заказу #{order.id}', 14)
    if used_parts:
        WarrantyCard.create_for_order(order, 'part', f'Гарантия на установленные запчасти по заказу #{order.id}', 30)
    db.session.commit()
    pdf_buffer = generate_act_pdf_buffer(order)
    send_order_ready_email_with_act(order, pdf_buffer)
    log_order_change(
        order_id=order.id,
        user_id=current_user.id,
        username=current_user.username,
        action='complete',
        comment=f'Заказ завершён, зарплата мастера {final_salary} руб.'
    )
    return jsonify({'status': 'ok', 'final_salary': final_salary})

@main_bp.route('/api/orders/<int:id>/mark-checked', methods=['POST'])
@role_required('admin', 'manager')
def mark_checked(id):
    order = Order.query.get_or_404(id)
    data = request.json
    order.is_checked = True
    order.checked_by = data.get('checked_by', current_user.full_name or current_user.username)
    order.checked_at = datetime.utcnow()
    checklist = data.get('checklist', {})
    order.checklist_data = json.dumps(checklist, ensure_ascii=False)
    db.session.commit()
    log_order_change(
        order_id=order.id,
        user_id=current_user.id,
        username=current_user.username,
        action='check',
        comment=f'Техника проверена {order.checked_by}, чек-лист: {checklist}'
    )
    return jsonify({'status': 'ok'})

@main_bp.route('/api/orders/<int:id>/wait-parts', methods=['POST'])
@role_required('admin', 'manager', 'master')
def toggle_wait_parts(id):
    order = Order.query.get_or_404(id)
    if order.status == 'completed':
        return jsonify({'error': 'Completed'}), 400
    order.status = 'waiting_parts' if order.status != 'waiting_parts' else 'in_progress'
    db.session.commit()
    log_order_change(
        order_id=order.id,
        user_id=current_user.id,
        username=current_user.username,
        action='wait_parts',
        comment=f'Статус изменён на {order.status}'
    )
    return jsonify({'status': 'ok', 'new_status': order.status})

@main_bp.route('/api/orders/<int:id>/logs', methods=['GET'])
@login_required
def get_order_logs(id):
    logs = OrderLog.query.filter_by(order_id=id).order_by(OrderLog.created_at.desc()).all()
    return jsonify([{
        'id': l.id,
        'username': l.username or 'Система',
        'action': l.action,
        'field_name': l.field_name,
        'old_value': l.old_value,
        'new_value': l.new_value,
        'comment': l.comment,
        'created_at': l.created_at.isoformat()
    } for l in logs])

# ------------------ Генерация PDF для скачивания ------------------
@main_bp.route('/api/orders/<int:id>/pdf-act')
@login_required
def generate_act_pdf(id):
    order = Order.query.get_or_404(id)
    if order.status != 'completed':
        flash('Акт можно сформировать только для завершённого заказа', 'warning')
        return redirect(url_for('main.orders_page'))
    pdf_buffer = generate_act_pdf_buffer(order)
    return send_file(pdf_buffer, as_attachment=True, download_name=f'act_order_{order.id}.pdf', mimetype='application/pdf')

# ------------------ API: Финансы ------------------
@main_bp.route('/api/finance/transactions', methods=['GET'])
@role_required('admin', 'accountant', 'manager')
def get_transactions():
    trans = FinanceTransaction.query.order_by(FinanceTransaction.date.desc()).all()
    return jsonify([{
        'id': t.id, 'type': t.type, 'category': t.category,
        'amount': t.amount, 'date': t.date.isoformat(), 'description': t.description
    } for t in trans])

@main_bp.route('/api/finance/balance', methods=['GET'])
@login_required
def get_balance():
    return jsonify({'balance': recalc_balance()})

@main_bp.route('/api/finance/transaction', methods=['POST'])
@role_required('admin', 'accountant', 'manager')
def add_finance():
    data = request.json
    add_transaction(data['type'], data['category'], data['amount'], data.get('description', ''))
    return jsonify({'status': 'ok', 'balance': recalc_balance()})

# ------------------ API: Ежемесячные платежи ------------------
@main_bp.route('/api/recurring', methods=['GET'])
@role_required('admin', 'accountant')
def get_recurring():
    items = RecurringPayment.query.all()
    return jsonify([{
        'id': i.id, 'category': i.category, 'amount': i.amount,
        'description': i.description, 'last_applied_month': i.last_applied_month
    } for i in items])

@main_bp.route('/api/recurring', methods=['POST'])
@role_required('admin', 'accountant')
def add_recurring():
    data = request.json
    rec = RecurringPayment(
        category=data['category'],
        amount=data['amount'],
        description=data.get('description', ''),
        last_applied_month=None
    )
    db.session.add(rec)
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/recurring/<int:id>', methods=['DELETE'])
@role_required('admin', 'accountant')
def delete_recurring(id):
    rec = RecurringPayment.query.get_or_404(id)
    db.session.delete(rec)
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/recurring/apply', methods=['POST'])
@role_required('admin', 'accountant')
def apply_recurring():
    now = datetime.utcnow()
    current_ym = now.strftime('%Y-%m')
    applied = 0
    for pay in RecurringPayment.query.all():
        if pay.last_applied_month != current_ym:
            add_transaction('expense', pay.category, pay.amount, pay.description or 'Ежемесячный платёж')
            pay.last_applied_month = current_ym
            applied += 1
    db.session.commit()
    return jsonify({'status': 'ok', 'applied': applied})

# ------------------ API: Чёрный список ------------------
@main_bp.route('/api/blacklist', methods=['GET'])
@role_required('admin', 'manager')
def get_blacklist():
    items = BlacklistClient.query.all()
    return jsonify([{'id': i.id, 'full_name': i.full_name, 'phone': i.phone, 'reason': i.reason} for i in items])

@main_bp.route('/api/blacklist', methods=['POST'])
@role_required('admin', 'manager')
def add_blacklist():
    data = request.json
    bl = BlacklistClient(full_name=data['full_name'], phone=data['phone'], reason=data.get('reason', ''))
    db.session.add(bl)
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/blacklist/<int:id>', methods=['DELETE'])
@role_required('admin', 'manager')
def del_blacklist(id):
    bl = BlacklistClient.query.get_or_404(id)
    db.session.delete(bl)
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Прайс ------------------
@main_bp.route('/api/pricelist', methods=['GET'])
@login_required
def get_pricelist():
    items = PriceItem.query.all()
    return jsonify([{'id': i.id, 'name': i.name, 'execution_time': i.execution_time, 'price': i.price} for i in items])

@main_bp.route('/api/pricelist', methods=['POST'])
@role_required('admin', 'accountant', 'manager')
def add_price():
    data = request.json
    price = PriceItem(name=data['name'], execution_time=data.get('execution_time', ''), price=data['price'])
    db.session.add(price)
    db.session.commit()
    return jsonify({'status': 'ok'})

@main_bp.route('/api/pricelist/<int:id>', methods=['DELETE'])
@role_required('admin', 'accountant', 'manager')
def del_price(id):
    price = PriceItem.query.get_or_404(id)
    db.session.delete(price)
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Календарь ------------------
@main_bp.route('/api/daily-orders', methods=['GET'])
@login_required
def daily_orders():
    date_str = request.args.get('date')
    if not date_str:
        return jsonify([])
    target = datetime.strptime(date_str, '%Y-%m-%d').date()
    orders = Order.query.filter(func.date(Order.start_time) == target).all()
    return jsonify([{
        'id': o.id, 'customer_name': o.customer_name, 'device_model': o.device_model,
        'price': o.price, 'status': o.status
    } for o in orders])

# ------------------ API: Гарантии ------------------
@main_bp.route('/api/warranty-cards', methods=['GET'])
@login_required
def get_warranty_cards():
    cards = WarrantyCard.query.all()
    return jsonify([{
        'id': c.id,
        'order_id': c.order_id,
        'warranty_type': c.warranty_type,
        'description': c.description,
        'valid_until': c.valid_until.isoformat(),
        'is_active': c.is_active,
        'created_at': c.created_at.isoformat()
    } for c in cards])

@main_bp.route('/api/warranty-cards/<int:id>/deactivate', methods=['POST'])
@role_required('admin', 'manager')
def deactivate_warranty(id):
    card = WarrantyCard.query.get_or_404(id)
    card.is_active = False
    db.session.commit()
    return jsonify({'status': 'ok'})

# ------------------ API: Графики ------------------
@main_bp.route('/api/chart-orders-status')
@login_required
def chart_orders_status():
    in_progress = Order.query.filter_by(status='in_progress').count()
    waiting = Order.query.filter_by(status='waiting_parts').count()
    completed = Order.query.filter_by(status='completed').count()
    return jsonify({
        'labels': ['В работе', 'Ожидают запчасть', 'Завершены'],
        'data': [in_progress, waiting, completed],
        'colors': ['#ffc107', '#0dcaf0', '#198754']
    })

@main_bp.route('/api/chart-revenue-daily')
@login_required
def chart_revenue_daily():
    today = datetime.utcnow().date()
    dates = []
    revenues = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        dates.append(day.strftime('%d.%m'))
        start = datetime(day.year, day.month, day.day)
        end = start + timedelta(days=1)
        total = db.session.query(func.sum(FinanceTransaction.amount)).filter(
            FinanceTransaction.type == 'income',
            FinanceTransaction.date >= start,
            FinanceTransaction.date < end
        ).scalar() or 0
        revenues.append(total)
    return jsonify({'labels': dates, 'data': revenues})

@main_bp.route('/api/chart-master-load')
@login_required
def chart_master_load():
    masters = Employee.query.filter_by(position='repair', fired=False).all()
    names = []
    counts = []
    for master in masters:
        active = Order.query.filter_by(responsible_employee_id=master.id).filter(Order.status != 'completed').count()
        names.append(master.full_name)
        counts.append(active)
    return jsonify({'labels': names, 'data': counts})

# ------------------ Сброс демо-данных ------------------
@main_bp.route('/api/reset-demo', methods=['POST'])
@role_required('admin')
def reset_demo():
    WarehouseItem.query.delete()
    Employee.query.delete()
    Order.query.delete()
    FinanceTransaction.query.delete()
    BlacklistClient.query.delete()
    PriceItem.query.delete()
    RecurringPayment.query.delete()
    Notification.query.delete()
    OrderLog.query.delete()
    WarrantyCard.query.delete()
    demo_warehouse = [
        WarehouseItem(name="Дисплей iPhone", quantity=5, weight="0.1кг", size="6.1", cost_price=1500),
        WarehouseItem(name="Аккумулятор", quantity=10, weight="50г", size="стандарт", cost_price=800)
    ]
    demo_employees = [
        Employee(full_name="Иван Петров", passport="4512 345678", details="счет 40817", position="repair", salary_value=67),
        Employee(full_name="Ольга Смирнова", passport="4512 111222", details="", position="reception", salary_value=30000)
    ]
    demo_pricelist = [
        PriceItem(name="Замена экрана", execution_time="2 часа", price=2000),
        PriceItem(name="Чистка от пыли", execution_time="1 день", price=800)
    ]
    demo_recurring = [
        RecurringPayment(category="Аренда помещения", amount=15000, description="Аренда"),
        RecurringPayment(category="Электроэнергия", amount=3500, description="Свет"),
        RecurringPayment(category="Водоснабжение", amount=1200, description="Вода"),
        RecurringPayment(category="Отопление", amount=2800, description="Отопление"),
        RecurringPayment(category="Кредитный платёж", amount=5000, description="Кредит")
    ]
    for item in demo_warehouse:
        db.session.add(item)
    for emp in demo_employees:
        db.session.add(emp)
    for price in demo_pricelist:
        db.session.add(price)
    for rec in demo_recurring:
        db.session.add(rec)
    deadline = datetime.utcnow() + timedelta(days=2)
    demo_order = Order(
        customer_name="Алексей", phone="+79111234567", device_model="Xiaomi Note",
        serial_number="SN123", imei="356789", main_problem="Не включается",
        detected_problem="Батарея", price=2500, deadline=deadline,
        start_time=datetime.utcnow() - timedelta(days=1), status="in_progress",
        responsible_employee_id=1, is_checked=False
    )
    db.session.add(demo_order)
    add_transaction('income', 'Прием заказа', 2500, 'Заказ №1 (демо)')
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Демо-данные восстановлены'})

# ------------------ Экспорт в Excel ------------------
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

@main_bp.route('/api/export/orders')
@login_required
def export_orders_excel():
    """Экспорт заказов в Excel с учётом текущих фильтров"""
    # Получаем параметры фильтрации (как в /api/orders)
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = Order.query
    if search:
        query = query.filter(
            db.or_(
                Order.customer_name.ilike(f'%{search}%'),
                Order.phone.ilike(f'%{search}%'),
                Order.device_model.ilike(f'%{search}%')
            )
        )
    if status:
        query = query.filter(Order.status == status)
    if date_from:
        query = query.filter(Order.start_time >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(Order.start_time <= datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
    
    orders = query.order_by(Order.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Заголовки
    headers = ['ID', 'Клиент', 'Телефон', 'Модель', 'Проблема', 'Обнаружено', 'Цена', 'Дедлайн', 'Дата приёма', 'Дата завершения', 'Статус', 'Проверено']
    ws.append(headers)
    # Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for o in orders:
        ws.append([
            o.id,
            o.customer_name,
            o.phone,
            o.device_model,
            o.main_problem,
            o.detected_problem,
            o.price,
            o.deadline.strftime('%d.%m.%Y %H:%M') if o.deadline else '',
            o.start_time.strftime('%d.%m.%Y %H:%M'),
            o.completed_at.strftime('%d.%m.%Y %H:%M') if o.completed_at else '',
            {'in_progress': 'В работе', 'waiting_parts': 'Ожидание запчасти', 'completed': 'Завершён'}.get(o.status, o.status),
            'Да' if o.is_checked else 'Нет'
        ])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'orders_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main_bp.route('/api/export/finance')
@role_required('admin', 'accountant', 'manager')
def export_finance_excel():
    """Экспорт финансовых транзакций"""
    trans = FinanceTransaction.query.order_by(FinanceTransaction.date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансы"
    headers = ['ID', 'Дата', 'Тип', 'Категория', 'Сумма', 'Описание']
    ws.append(headers)
    for t in trans:
        ws.append([
            t.id,
            t.date.strftime('%d.%m.%Y %H:%M'),
            'Доход' if t.type == 'income' else 'Расход',
            t.category,
            t.amount,
            t.description or ''
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'finance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main_bp.route('/api/export/employees')
@role_required('admin', 'manager')
def export_employees_excel():
    """Экспорт сотрудников"""
    emps = Employee.query.filter_by(fired=False).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    headers = ['ID', 'ФИО', 'Паспорт', 'Реквизиты', 'Должность', 'ЗП (ставка/%)']
    ws.append(headers)
    for e in emps:
        position_map = {'repair': 'Ремонт', 'reception': 'Приёмка', 'cleaning': 'Уборка'}
        ws.append([e.id, e.full_name, e.passport, e.details, position_map.get(e.position, e.position), e.salary_value])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'employees_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main_bp.route('/api/chart-orders-daily')
@login_required
def chart_orders_daily():
    """Количество заказов по дням за последние 7 дней"""
    today = datetime.utcnow().date()
    dates = []
    counts = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        dates.append(day.strftime('%d.%m'))
        start = datetime(day.year, day.month, day.day)
        end = start + timedelta(days=1)
        count = Order.query.filter(Order.start_time >= start, Order.start_time < end).count()
        counts.append(count)
    return jsonify({'labels': dates, 'data': counts})

@main_bp.route('/api/chart-top-masters')
@login_required
def chart_top_masters():
    """Топ-5 мастеров по количеству завершённых заказов"""
    masters = Employee.query.filter_by(position='repair', fired=False).all()
    stats = []
    for master in masters:
        completed_count = Order.query.filter_by(responsible_employee_id=master.id, status='completed').count()
        stats.append({'name': master.full_name, 'count': completed_count})
    stats.sort(key=lambda x: x['count'], reverse=True)
    top5 = stats[:5]
    return jsonify({'labels': [s['name'] for s in top5], 'data': [s['count'] for s in top5]})

@main_bp.route('/api/chart-popular-models')
@login_required
def chart_popular_models():
    """Топ-5 моделей устройств по количеству заказов"""
    from sqlalchemy import func
    popular = db.session.query(Order.device_model, func.count(Order.id).label('cnt')).filter(Order.device_model != '').group_by(Order.device_model).order_by(func.count(Order.id).desc()).limit(5).all()
    return jsonify({'labels': [p[0] for p in popular], 'data': [p[1] for p in popular]})

@main_bp.route('/api/urgent-orders')
@login_required
def urgent_orders():
    """Заказы с дедлайном сегодня или завтра (не завершённые)"""
    now = datetime.utcnow()
    today_end = datetime(now.year, now.month, now.day, 23, 59, 59)
    tomorrow_end = today_end + timedelta(days=1)
    orders = Order.query.filter(Order.deadline <= tomorrow_end, Order.deadline >= now, Order.status != 'completed').order_by(Order.deadline).limit(10).all()
    return jsonify([{
        'id': o.id, 'customer_name': o.customer_name, 'device_model': o.device_model,
        'deadline': o.deadline.strftime('%d.%m.%Y %H:%M'), 'status': o.status
    } for o in orders])

# ------------------ Массовые операции с заказами ------------------
@main_bp.route('/api/orders/bulk', methods=['POST'])
@role_required('admin', 'manager')
def bulk_order_action():
    data = request.json
    order_ids = data.get('order_ids', [])
    action = data.get('action')  # 'status', 'assign_master', 'delete'
    value = data.get('value')
    
    if not order_ids:
        return jsonify({'error': 'Нет выбранных заказов'}), 400
    
    orders = Order.query.filter(Order.id.in_(order_ids)).all()
    
    if action == 'status':
        if value not in ['in_progress', 'waiting_parts', 'completed']:
            return jsonify({'error': 'Неверный статус'}), 400
        for order in orders:
            old_status = order.status
            order.status = value
            log_order_change(
                order_id=order.id,
                user_id=current_user.id,
                username=current_user.username,
                action='edit',
                comment=f'Массовое изменение статуса: {old_status} → {value}'
            )
    elif action == 'assign_master':
        master = Employee.query.get(value)
        if not master or master.position != 'repair':
            return jsonify({'error': 'Неверный мастер'}), 400
        for order in orders:
            old_master = order.responsible_employee_id
            order.responsible_employee_id = master.id
            log_order_change(
                order_id=order.id,
                user_id=current_user.id,
                username=current_user.username,
                action='edit',
                comment=f'Массовое назначение мастера: #{old_master} → #{master.id}'
            )
    elif action == 'delete':
        for order in orders:
            log_order_change(
                order_id=order.id,
                user_id=current_user.id,
                username=current_user.username,
                action='edit',
                comment='Массовое удаление заказа'
            )
            db.session.delete(order)
    else:
        return jsonify({'error': 'Неизвестное действие'}), 400
    
    db.session.commit()
    return jsonify({'status': 'ok', 'updated': len(orders)})

@main_bp.route('/api/export/orders-selected', methods=['POST'])
@login_required
def export_selected_orders_excel():
    data = request.json
    order_ids = data.get('order_ids', [])
    if not order_ids:
        return jsonify({'error': 'Нет выбранных заказов'}), 400
    
    orders = Order.query.filter(Order.id.in_(order_ids)).order_by(Order.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Выбранные заказы"
    headers = ['ID', 'Клиент', 'Телефон', 'Модель', 'Проблема', 'Обнаружено', 'Цена', 'Дедлайн', 'Дата приёма', 'Дата завершения', 'Статус', 'Проверено']
    ws.append(headers)
    for o in orders:
        ws.append([
            o.id, o.customer_name, o.phone, o.device_model, o.main_problem,
            o.detected_problem, o.price,
            o.deadline.strftime('%d.%m.%Y %H:%M') if o.deadline else '',
            o.start_time.strftime('%d.%m.%Y %H:%M'),
            o.completed_at.strftime('%d.%m.%Y %H:%M') if o.completed_at else '',
            {'in_progress': 'В работе', 'waiting_parts': 'Ожидание запчасти', 'completed': 'Завершён'}.get(o.status, o.status),
            'Да' if o.is_checked else 'Нет'
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'selected_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')